import openpyxl as excel

#data_only permite obtener los datos vistos en el libro
fileData = excel.load_workbook("excel_files/dataPeople.xlsx", data_only=True)

fileData_sheet = fileData["Hoja1"]

kids = {}
adults = {}
marriedState = []

for row in range(2, fileData_sheet.max_row):
  
    #creando direcciones de celdas para obtener sus valores
    name_cell = "A" + str(row)
    age_cell = "B" + str(row)
    married_cell = "D" + str(row)

    name = fileData_sheet[name_cell].value
    age = fileData_sheet[age_cell].value
    married = fileData_sheet[married_cell].value

    if name or age or married != None:

        #which ones are kids and adults, and total of them
        if age < 18:
            kids[name] = age
        else:
            adults[name] = age

        #is there anyone married?
        if married == "Yes":
            marriedState.append(name)

    else:
         break

kidsTotal = len(kids)
adultsTotal = len(adults)

#imprimiendo los resultados en otra hoja

# La hoja ya existe?
if "Results" in fileData.sheetnames:
    results_sheet = fileData["Results"]
else:
    fileData.create_sheet("Results")
    results_sheet = fileData["Results"]

# imprimiendo datos
def DataPrinting(forWho):

    count = 3

    if forWho == "kids":
        results_sheet.merge_cells("A1:B1")
        results_sheet["A1"] = "Kids list"

        results_sheet["A2"] = "Name"
        results_sheet["B2"] = "Age"

        for kid in kids:
            results_sheet["A" + str(count)] = kid
            results_sheet["B" + str(count)] = kids[kid]
            count += 1
        print("ok")

    elif forWho == "adults":
        results_sheet.merge_cells("D1:E1")
        results_sheet["D1"] = "Adult list"

        results_sheet["D2"] = "Name"
        results_sheet["E2"] = "Age"

        for adult in adults:
            results_sheet["D" + str(count)] = adult
            results_sheet["E" + str(count)] = adults[adult]
            count += 1
        print("ok")

    elif forWho == "married":

        results_sheet["G1"] = "Married list"
        results_sheet["G2"] = "Name"

        for marriedP in marriedState:
            results_sheet["G" + str(count)] = marriedP
            count += 1
        print("ok")
    
    else:
        print("Dato desconocido")

DataPrinting("kids")
DataPrinting("adults")
DataPrinting("married")

fileData.save("excel_files/dataPeople.xlsx")

#print(f"Kids = {kids}\nAdult = {adults}\n Kids total: {kidsTotal} - adults total: {adultsTotal} \n married: {marriedState}")